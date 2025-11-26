from rest_framework import status
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from django.core.management import call_command
from django.conf import settings
from django.db import connection
from django.http import FileResponse, HttpResponse
import os
import json
from datetime import datetime
from apps.logs.utils import create_log, LogTimer
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_database_stats(request):
    """
    Récupérer les statistiques de la base de données (admin uniquement)
    """
    try:
        # Vérifier que l'utilisateur est authentifié
        if not request.user or not request.user.is_authenticated:
            return Response({
                'error': 'Authentification requise'
            }, status=status.HTTP_401_UNAUTHORIZED)
        
        # Vérifier que l'utilisateur a un rôle
        if not hasattr(request.user, 'role'):
            print(f"ERREUR: L'utilisateur {request.user} n'a pas d'attribut 'role'")
            return Response({
                'error': 'Utilisateur invalide'
            }, status=status.HTTP_403_FORBIDDEN)
        
        # Seuls les admins peuvent voir
        if request.user.role != 'admin':
            return Response({
                'error': 'Permission refusée. Rôle admin requis.'
            }, status=status.HTTP_403_FORBIDDEN)
    
    except AttributeError as e:
        print(f"AttributeError dans get_database_stats: {e}")
        return Response({
            'error': f'Erreur d\'attribut: {str(e)}'
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
    try:
        from apps.authentication.models import User
        from apps.clients.models import Client
        from apps.products.models import Produit, MouvementStock
        from apps.orders.models import Commande, ItemCommande
        from apps.logs.models import SystemLog
        
        # Compter les enregistrements
        stats = {
            'users': User.objects.count(),
            'clients': Client.objects.count(),
            'products': Produit.objects.count(),
            'stock_movements': MouvementStock.objects.count(),
            'orders': Commande.objects.count(),
            'order_lines': ItemCommande.objects.count(),
            'deliveries': Commande.objects.filter(statut='livree').count(),
            'logs': SystemLog.objects.count(),
        }
        
        # Taille de la base de données (si SQLite)
        try:
            db_path = settings.DATABASES['default']['NAME']
            # Convertir Path en string si nécessaire
            db_path_str = str(db_path) if db_path else None
            if db_path_str and os.path.exists(db_path_str):
                db_size_bytes = os.path.getsize(db_path_str)
                db_size_mb = round(db_size_bytes / (1024 * 1024), 2)
                stats['database_size_mb'] = db_size_mb
            else:
                stats['database_size_mb'] = 0
        except Exception as e:
            print(f"Erreur lors de la récupération de la taille de la DB: {e}")
            stats['database_size_mb'] = 0
        
        # Total des enregistrements
        stats['total_records'] = sum([v for k, v in stats.items() if k != 'database_size_mb'])
        
        # Dernière sauvegarde
        backup_dir = os.path.join(settings.BASE_DIR, 'backups')
        if os.path.exists(backup_dir):
            backups = [f for f in os.listdir(backup_dir) if f.endswith('.json')]
            if backups:
                latest_backup = max(backups)
                stats['last_backup'] = latest_backup
                backup_path = os.path.join(backup_dir, latest_backup)
                backup_size = os.path.getsize(backup_path)
                stats['last_backup_size_mb'] = round(backup_size / (1024 * 1024), 2)
            else:
                stats['last_backup'] = None
        
        return Response(stats)
        
    except Exception as e:
        import traceback
        print(f"Erreur dans get_database_stats: {str(e)}")
        print(traceback.format_exc())
        return Response({
            'error': str(e),
            'details': traceback.format_exc()
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def create_backup(request):
    """
    Créer une sauvegarde complète de la base de données (admin uniquement)
    Génère à la fois un fichier JSON et un fichier Excel
    """
    # Seuls les admins peuvent créer des sauvegardes
    if request.user.role != 'admin':
        return Response({
            'error': 'Permission refusée'
        }, status=status.HTTP_403_FORBIDDEN)
    
    with LogTimer() as timer:
        try:
            # Créer le dossier de sauvegarde s'il n'existe pas
            backup_dir = os.path.join(settings.BASE_DIR, 'backups')
            os.makedirs(backup_dir, exist_ok=True)
            
            # Nom des fichiers de sauvegarde
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            json_filename = f'backup_{timestamp}.json'
            excel_filename = f'backup_{timestamp}.xlsx'
            json_path = os.path.join(backup_dir, json_filename)
            excel_path = os.path.join(backup_dir, excel_filename)
            
            # 1. Créer la sauvegarde JSON avec dumpdata
            with open(json_path, 'w', encoding='utf-8') as f:
                call_command('dumpdata', 
                           exclude=['contenttypes', 'auth.permission', 'sessions'],
                           indent=2,
                           stdout=f)
            
            # 2. Créer la sauvegarde Excel
            from apps.authentication.models import User
            from apps.clients.models import Client
            from apps.products.models import Produit, MouvementStock
            from apps.orders.models import Commande, ItemCommande
            from apps.logs.models import SystemLog
            
            wb = openpyxl.Workbook()
            wb.remove(wb.active)  # Supprimer la feuille par défaut
            
            # Style pour les en-têtes
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="0066CC", end_color="0066CC", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")
            
            # Feuille 1: Utilisateurs
            ws_users = wb.create_sheet("Utilisateurs")
            users_headers = ['ID', 'Email', 'Prénom', 'Nom', 'Rôle', 'Téléphone', 'Actif', 'Date création']
            ws_users.append(users_headers)
            for cell in ws_users[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
            
            for user in User.objects.all():
                ws_users.append([
                    user.id,
                    user.email,
                    user.first_name,
                    user.last_name,
                    user.role,
                    user.telephone or '',
                    'Oui' if user.is_active else 'Non',
                    user.date_creation.strftime('%d/%m/%Y %H:%M') if user.date_creation else ''
                ])
            
            # Ajuster la largeur des colonnes
            for col in range(1, len(users_headers) + 1):
                ws_users.column_dimensions[get_column_letter(col)].width = 15
            
            # Feuille 2: Clients
            ws_clients = wb.create_sheet("Clients")
            clients_headers = ['ID', 'Nom commercial', 'Raison sociale', 'Contact', 'Téléphone', 'Email', 'Adresse']
            ws_clients.append(clients_headers)
            for cell in ws_clients[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
            
            for client in Client.objects.all():
                ws_clients.append([
                    client.id,
                    client.nom_commercial or '',
                    client.raison_sociale or '',
                    client.contact or '',
                    client.telephone or '',
                    client.email or '',
                    client.adresse or ''
                ])
            
            for col in range(1, len(clients_headers) + 1):
                ws_clients.column_dimensions[get_column_letter(col)].width = 20
            
            # Feuille 3: Produits
            ws_products = wb.create_sheet("Produits")
            products_headers = ['ID', 'Nom', 'Type', 'Prix unitaire', 'Stock', 'Stock min', 'Actif']
            ws_products.append(products_headers)
            for cell in ws_products[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
            
            for product in Produit.objects.all():
                ws_products.append([
                    product.id,
                    product.nom,
                    product.type_produit,
                    float(product.prix_unitaire),
                    product.stock_actuel,
                    product.stock_minimal,
                    'Oui' if product.is_active else 'Non'
                ])
            
            for col in range(1, len(products_headers) + 1):
                ws_products.column_dimensions[get_column_letter(col)].width = 15
            
            # Feuille 4: Commandes
            ws_orders = wb.create_sheet("Commandes")
            orders_headers = ['ID', 'Numéro', 'Client', 'Date', 'Statut', 'Montant produits', 'Frais livraison', 'Montant total', 'Créé par']
            ws_orders.append(orders_headers)
            for cell in ws_orders[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
            
            for order in Commande.objects.select_related('client', 'vendeur').all():
                ws_orders.append([
                    order.id,
                    order.numero_commande,
                    order.client.nom_commercial if order.client else '',
                    order.date_creation.strftime('%d/%m/%Y'),
                    order.statut,
                    float(order.montant_produits),
                    float(order.frais_livraison),
                    float(order.montant_total),
                    order.vendeur.email if order.vendeur else ''
                ])
            
            for col in range(1, len(orders_headers) + 1):
                ws_orders.column_dimensions[get_column_letter(col)].width = 18
            
            # Feuille 5: Mouvements de stock
            ws_stock = wb.create_sheet("Mouvements Stock")
            stock_headers = ['ID', 'Produit', 'Type', 'Quantité', 'Date', 'Motif', 'Utilisateur']
            ws_stock.append(stock_headers)
            for cell in ws_stock[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
            
            for mvt in MouvementStock.objects.select_related('produit', 'utilisateur').all():
                ws_stock.append([
                    mvt.id,
                    mvt.produit.nom if mvt.produit else '',
                    mvt.type_mouvement,
                    mvt.quantite,
                    mvt.date_creation.strftime('%d/%m/%Y %H:%M'),
                    mvt.motif or '',
                    mvt.utilisateur.email if mvt.utilisateur else ''
                ])
            
            for col in range(1, len(stock_headers) + 1):
                ws_stock.column_dimensions[get_column_letter(col)].width = 20
            
            # Feuille 6: Logs système
            ws_logs = wb.create_sheet("Logs Système")
            logs_headers = ['ID', 'Type', 'Module', 'Message', 'Utilisateur', 'Date', 'IP']
            ws_logs.append(logs_headers)
            for cell in ws_logs[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
            
            for log in SystemLog.objects.select_related('user').order_by('-timestamp')[:1000]:  # Limiter à 1000 logs
                ws_logs.append([
                    log.id,
                    log.type,
                    log.module,
                    log.message,
                    log.user.email if log.user else 'Système',
                    log.timestamp.strftime('%d/%m/%Y %H:%M:%S'),
                    log.ip_address or ''
                ])
            
            for col in range(1, len(logs_headers) + 1):
                ws_logs.column_dimensions[get_column_letter(col)].width = 25
            
            # Sauvegarder le fichier Excel
            wb.save(excel_path)
            
            # Taille des fichiers
            json_size = os.path.getsize(json_path)
            json_size_mb = round(json_size / (1024 * 1024), 2)
            excel_size = os.path.getsize(excel_path)
            excel_size_mb = round(excel_size / (1024 * 1024), 2)
            total_size_mb = round(json_size_mb + excel_size_mb, 2)
            
            # Créer un log
            try:
                create_log(
                    log_type='success',
                    message="Sauvegarde de la base de données créée (JSON + Excel)",
                    details=f"JSON: {json_filename} ({json_size_mb} MB), Excel: {excel_filename} ({excel_size_mb} MB)",
                    user=request.user,
                    module='database',
                    request=request,
                    metadata={
                        'json_filename': json_filename,
                        'excel_filename': excel_filename,
                        'json_size_mb': json_size_mb,
                        'excel_size_mb': excel_size_mb,
                        'timestamp': timestamp
                    },
                    status_code=200,
                    response_time=timer.elapsed
                )
            except Exception:
                pass
            
            return Response({
                'message': 'Sauvegarde créée avec succès',
                'json_filename': json_filename,
                'excel_filename': excel_filename,
                'json_size_mb': json_size_mb,
                'excel_size_mb': excel_size_mb,
                'total_size_mb': total_size_mb,
                'timestamp': timestamp
            })
            
        except Exception as e:
            # Créer un log d'erreur
            try:
                create_log(
                    log_type='error',
                    message="Erreur lors de la création de la sauvegarde",
                    details=str(e),
                    user=request.user,
                    module='database',
                    request=request,
                    metadata={'error': str(e)},
                    status_code=500,
                    response_time=timer.elapsed
                )
            except Exception:
                pass
            
            return Response({
                'error': f'Erreur lors de la sauvegarde: {str(e)}'
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def list_backups(request):
    """
    Lister toutes les sauvegardes disponibles (admin uniquement)
    Inclut les fichiers JSON et Excel
    """
    # Seuls les admins peuvent voir
    if request.user.role != 'admin':
        return Response({
            'error': 'Permission refusée'
        }, status=status.HTTP_403_FORBIDDEN)
    
    try:
        backup_dir = os.path.join(settings.BASE_DIR, 'backups')
        
        if not os.path.exists(backup_dir):
            return Response({'backups': []})
        
        # Grouper les sauvegardes par timestamp
        backups_dict = {}
        
        for filename in os.listdir(backup_dir):
            if filename.startswith('backup_') and (filename.endswith('.json') or filename.endswith('.xlsx')):
                # Extraire le timestamp du nom de fichier
                timestamp = filename.replace('backup_', '').replace('.json', '').replace('.xlsx', '')
                
                if timestamp not in backups_dict:
                    backups_dict[timestamp] = {
                        'timestamp': timestamp,
                        'json_file': None,
                        'excel_file': None,
                        'json_size_mb': 0,
                        'excel_size_mb': 0,
                        'created': None,
                        'created_formatted': None
                    }
                
                filepath = os.path.join(backup_dir, filename)
                size_bytes = os.path.getsize(filepath)
                size_mb = round(size_bytes / (1024 * 1024), 2)
                created = datetime.fromtimestamp(os.path.getctime(filepath))
                
                if filename.endswith('.json'):
                    backups_dict[timestamp]['json_file'] = filename
                    backups_dict[timestamp]['json_size_mb'] = size_mb
                elif filename.endswith('.xlsx'):
                    backups_dict[timestamp]['excel_file'] = filename
                    backups_dict[timestamp]['excel_size_mb'] = size_mb
                
                backups_dict[timestamp]['created'] = created.isoformat()
                backups_dict[timestamp]['created_formatted'] = created.strftime('%d/%m/%Y à %H:%M')
        
        # Convertir en liste et calculer la taille totale
        backups = []
        for timestamp, data in backups_dict.items():
            data['total_size_mb'] = round(data['json_size_mb'] + data['excel_size_mb'], 2)
            backups.append(data)
        
        # Trier par date de création (plus récent en premier)
        backups.sort(key=lambda x: x['created'] if x['created'] else '', reverse=True)
        
        return Response({'backups': backups})
        
    except Exception as e:
        return Response({
            'error': str(e)
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def restore_backup(request):
    """
    Restaurer une sauvegarde (admin uniquement)
    ATTENTION: Cette opération remplace toutes les données actuelles
    """
    # Seuls les admins peuvent restaurer
    if request.user.role != 'admin':
        return Response({
            'error': 'Permission refusée'
        }, status=status.HTTP_403_FORBIDDEN)
    
    filename = request.data.get('filename')
    if not filename:
        return Response({
            'error': 'Nom de fichier requis'
        }, status=status.HTTP_400_BAD_REQUEST)
    
    with LogTimer() as timer:
        try:
            backup_dir = os.path.join(settings.BASE_DIR, 'backups')
            backup_path = os.path.join(backup_dir, filename)
            
            if not os.path.exists(backup_path):
                return Response({
                    'error': 'Fichier de sauvegarde introuvable'
                }, status=status.HTTP_404_NOT_FOUND)
            
            # Utiliser loaddata pour restaurer
            call_command('loaddata', backup_path)
            
            # Créer un log
            try:
                create_log(
                    log_type='warning',
                    message="Base de données restaurée depuis une sauvegarde",
                    details=f"Fichier: {filename}",
                    user=request.user,
                    module='database',
                    request=request,
                    metadata={
                        'backup_filename': filename,
                        'restored_at': datetime.now().isoformat()
                    },
                    status_code=200,
                    response_time=timer.elapsed
                )
            except Exception:
                pass
            
            return Response({
                'message': 'Base de données restaurée avec succès',
                'filename': filename
            })
            
        except Exception as e:
            # Créer un log d'erreur
            try:
                create_log(
                    log_type='error',
                    message="Erreur lors de la restauration",
                    details=str(e),
                    user=request.user,
                    module='database',
                    request=request,
                    metadata={'error': str(e), 'filename': filename},
                    status_code=500,
                    response_time=timer.elapsed
                )
            except Exception:
                pass
            
            return Response({
                'error': f'Erreur lors de la restauration: {str(e)}'
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def download_backup(request, filename):
    """
    Télécharger un fichier de sauvegarde (JSON ou Excel)
    """
    # Seuls les admins peuvent télécharger
    if request.user.role != 'admin':
        return Response({
            'error': 'Permission refusée'
        }, status=status.HTTP_403_FORBIDDEN)
    
    try:
        backup_dir = os.path.join(settings.BASE_DIR, 'backups')
        file_path = os.path.join(backup_dir, filename)
        
        # Vérifier que le fichier existe
        if not os.path.exists(file_path):
            return Response({
                'error': 'Fichier introuvable'
            }, status=status.HTTP_404_NOT_FOUND)
        
        # Vérifier que c'est bien un fichier de sauvegarde
        if not (filename.startswith('backup_') and (filename.endswith('.json') or filename.endswith('.xlsx'))):
            return Response({
                'error': 'Fichier non autorisé'
            }, status=status.HTTP_403_FORBIDDEN)
        
        # Déterminer le type de contenu
        if filename.endswith('.json'):
            content_type = 'application/json'
        elif filename.endswith('.xlsx'):
            content_type = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        else:
            content_type = 'application/octet-stream'
        
        # Ouvrir et retourner le fichier
        response = FileResponse(open(file_path, 'rb'), content_type=content_type)
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        # Logger le téléchargement
        try:
            create_log(
                log_type='info',
                message=f"Téléchargement de sauvegarde: {filename}",
                details=f"Utilisateur: {request.user.email}",
                user=request.user,
                module='database',
                request=request,
                metadata={'filename': filename},
                status_code=200,
                response_time=0
            )
        except Exception:
            pass
        
        return response
        
    except Exception as e:
        return Response({
            'error': f'Erreur lors du téléchargement: {str(e)}'
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
